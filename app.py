"""
OSAM SEO Rank Refresher
=======================
Streamlit app that connects to OneDrive via Microsoft Graph API,
reads the OSAM SEO Tracking Excel file, calculates rank metrics
for all project sheets, and writes results back to the Rank Summary sheet.

One-time setup: configure secrets in Streamlit Cloud.
After that: single URL, single click, works on any browser/OS.
"""

import streamlit as st
import requests
import msal
import math
import re
import json
from datetime import date, datetime
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="OSAM SEO Rank Refresher",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ─────────────────────────────────────────────
# CUSTOM CSS — Clean professional dark UI
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

.stApp {
    background: #0f1117;
    color: #e8eaf0;
}

/* Header */
.app-header {
    background: linear-gradient(135deg, #1a1f2e 0%, #0f1117 100%);
    border: 1px solid #2a2f3e;
    border-radius: 16px;
    padding: 32px 36px;
    margin-bottom: 28px;
    display: flex;
    align-items: center;
    gap: 20px;
}
.header-icon {
    font-size: 40px;
    line-height: 1;
}
.header-title {
    font-size: 26px;
    font-weight: 700;
    color: #ffffff;
    letter-spacing: -0.5px;
    margin: 0;
}
.header-sub {
    font-size: 13px;
    color: #6b7280;
    margin: 4px 0 0;
    font-weight: 400;
}

/* Cards */
.card {
    background: #1a1f2e;
    border: 1px solid #2a2f3e;
    border-radius: 12px;
    padding: 24px;
    margin-bottom: 20px;
}
.card-title {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: #4b9fff;
    margin-bottom: 16px;
}

/* Status messages */
.status-box {
    border-radius: 10px;
    padding: 14px 18px;
    margin: 12px 0;
    font-size: 13.5px;
    font-weight: 500;
    display: flex;
    align-items: flex-start;
    gap: 10px;
    line-height: 1.5;
}
.status-info  { background: #1e2a3e; border: 1px solid #2a4a7f; color: #7eb8ff; }
.status-ok    { background: #1a2e1e; border: 1px solid #2a5a30; color: #6ddb7a; }
.status-warn  { background: #2e2a1a; border: 1px solid #5a4a20; color: #dbb86d; }
.status-error { background: #2e1a1a; border: 1px solid #5a2a2a; color: #db6d6d; }

/* Metric grid */
.metric-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 12px;
    margin-top: 16px;
}
.metric-card {
    background: #0f1117;
    border: 1px solid #2a2f3e;
    border-radius: 10px;
    padding: 16px;
    text-align: center;
}
.metric-value {
    font-size: 32px;
    font-weight: 700;
    font-family: 'DM Mono', monospace;
    line-height: 1;
    margin-bottom: 6px;
}
.metric-label {
    font-size: 11px;
    color: #6b7280;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    font-weight: 500;
}

/* Progress */
.progress-row {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 8px 0;
    border-bottom: 1px solid #1e2332;
    font-size: 13px;
}
.progress-row:last-child { border-bottom: none; }

/* Streamlit overrides */
div[data-testid="stSelectbox"] label,
div[data-testid="stButton"] { margin-top: 8px; }

.stSelectbox > div > div {
    background: #1a1f2e !important;
    border: 1px solid #2a2f3e !important;
    color: #e8eaf0 !important;
    border-radius: 8px !important;
}

.stButton > button {
    width: 100%;
    background: linear-gradient(135deg, #2563eb, #1d4ed8) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 14px 24px !important;
    font-size: 15px !important;
    font-weight: 600 !important;
    font-family: 'DM Sans', sans-serif !important;
    letter-spacing: 0.02em !important;
    transition: all 0.2s !important;
    cursor: pointer !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #1d4ed8, #1e40af) !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 8px 24px rgba(37, 99, 235, 0.35) !important;
}
.stButton > button:disabled {
    background: #2a2f3e !important;
    color: #4b5563 !important;
    cursor: not-allowed !important;
}

div[data-testid="stExpander"] {
    background: #1a1f2e !important;
    border: 1px solid #2a2f3e !important;
    border-radius: 10px !important;
}

.streamlit-expanderHeader { color: #9ca3af !important; font-size: 13px !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# CONFIGURATION — edit these once
# ─────────────────────────────────────────────
# These are read from Streamlit secrets (secrets.toml)
# In secrets.toml:
#   [onedrive]
#   client_id     = "YOUR_CLIENT_ID"       # from Azure (free, no credit card for personal)
#   client_secret = "YOUR_CLIENT_SECRET"
#   tenant_id     = "consumers"            # for personal accounts always use "consumers"
#   file_res_id   = "627132b3-3264-43df-bf65-f22824cdf67e"
#   user_email    = "your@outlook.com"

def get_config():
    try:
        return {
            "client_id":     st.secrets["onedrive"]["client_id"],
            "client_secret": st.secrets["onedrive"]["client_secret"],
            "tenant_id":     st.secrets["onedrive"]["tenant_id"],
            "file_res_id":   st.secrets["onedrive"]["file_res_id"],
            "user_email":    st.secrets["onedrive"]["user_email"],
        }
    except Exception:
        # Fallback for local testing — fill these in locally
        return {
            "client_id":     "YOUR_CLIENT_ID",
            "client_secret": "YOUR_CLIENT_SECRET",
            "tenant_id":     "consumers",
            "file_res_id":   "627132b3-3264-43df-bf65-f22824cdf67e",
            "user_email":    "YOUR_EMAIL@outlook.com",
        }


# ─────────────────────────────────────────────
# MICROSOFT GRAPH — AUTH & FILE ACCESS
# ─────────────────────────────────────────────

@st.cache_data(ttl=3000, show_spinner=False)
def get_access_token(client_id, client_secret, tenant_id):
    """Get OAuth2 token using client credentials flow."""
    authority = f"https://login.microsoftonline.com/{tenant_id}"
    app = msal.ConfidentialClientApplication(
        client_id,
        authority=authority,
        client_credential=client_secret,
    )
    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" in result:
        return result["access_token"]
    raise Exception(f"Auth failed: {result.get('error_description', result.get('error', 'Unknown error'))}")


def graph_get(token, url):
    """Make a GET request to Microsoft Graph API."""
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.json()


def download_workbook(token, file_res_id):
    """Download the Excel file as bytes via Graph API."""
    # Get file download URL using the item ID
    url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file_res_id}/content"
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers, timeout=60, allow_redirects=True)
    resp.raise_for_status()
    return BytesIO(resp.content)


def upload_workbook(token, file_res_id, workbook_bytes):
    """Upload the modified Excel file back to OneDrive."""
    url = f"https://graph.microsoft.com/v1.0/me/drive/items/{file_res_id}/content"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    resp = requests.put(url, headers=headers, data=workbook_bytes, timeout=120)
    resp.raise_for_status()
    return resp.status_code


# ─────────────────────────────────────────────
# RANK LOGIC — identical to Google Apps Script
# ─────────────────────────────────────────────

def parse_rank(val):
    """Convert any cell value to a rank number (1-100) or 'NA'."""
    if val is None:
        return "NA"
    s = str(val).strip().upper()
    if s in ("", "NA", "NULL"):
        return "NA"
    if not re.match(r'^\d+$', s):
        return "NA"
    n = int(s)
    if n < 1 or n > 100:
        return "NA"
    return n


def get_group(rank):
    """Return decade group (1-10=G1, 11-20=G2, etc.)"""
    if rank == "NA":
        return None
    return math.ceil(rank / 10)


def get_rank_status(p1, p2):
    """Return 'up', 'down', 'same', or 'na'."""
    if p1 == "NA" and p2 == "NA":
        return "na"
    if p1 == "NA" and p2 != "NA":
        return "up"      # newly appeared
    if p1 != "NA" and p2 == "NA":
        return "down"    # disappeared
    g1 = get_group(p1)
    g2 = get_group(p2)
    if g1 == g2:
        return "same"
    return "up" if p2 < p1 else "down"


def calculate_metrics(keywords):
    """Calculate all 6 metrics from a list of {prev, curr} dicts."""
    total = first_page = rank_up = rank_down = top5 = top3 = 0
    for kw in keywords:
        p1 = parse_rank(kw["prev"])
        p2 = parse_rank(kw["curr"])
        if p1 == "NA" and p2 == "NA":
            continue
        total += 1
        if p2 != "NA":
            if p2 <= 10: first_page += 1
            if p2 <= 5:  top5 += 1
            if p2 <= 3:  top3 += 1
        s = get_rank_status(p1, p2)
        if s == "up":   rank_up += 1
        if s == "down": rank_down += 1
    return {
        "total":       total,
        "first_page":  first_page,
        "rank_up":     rank_up,
        "rank_down":   rank_down,
        "top5":        top5,
        "top3":        top3,
    }


# ─────────────────────────────────────────────
# DATE LOGIC
# ─────────────────────────────────────────────

MONTH_MAP = {
    "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
    "jan":1,"feb":2,"mar":3,"apr":4,"jun":6,"jul":7,"aug":8,
    "sep":9,"sept":9,"oct":10,"nov":11,"dec":12
}

def parse_date_from_sheet_name(name):
    """Extract date from sheet name like 'March 31st Rank Summary'."""
    lower = name.lower()
    month_num = None
    for m, n in sorted(MONTH_MAP.items(), key=lambda x: -len(x[0])):
        if re.search(r'\b' + m + r'\b', lower):
            month_num = n
            break
    if not month_num:
        return None
    day_match = re.search(r'\b(\d{1,2})(st|nd|rd|th)?\b', lower)
    if not day_match:
        return None
    day = int(day_match.group(1))
    if day < 1 or day > 31:
        return None
    year = date.today().year
    try:
        return date(year, month_num, day)
    except ValueError:
        return None


def parse_cell_date(val):
    """Parse a date from a cell value — handles Date objects, DD/MM/YYYY, etc."""
    if val is None or val == "":
        return None

    # Already a date/datetime
    if isinstance(val, (date, datetime)):
        if isinstance(val, datetime):
            return val.date()
        return val

    s = str(val).strip()
    if not s:
        return None

    # Handle multiline cells — take last line
    if "\n" in s:
        s = s.split("\n")[-1].strip()

    # DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY
    m = re.match(r'^(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})$', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100: y += 2000
        if 1 <= mo <= 12 and 1 <= d <= 31:
            try:
                return date(y, mo, d)
            except ValueError:
                pass

    # YYYY-MM-DD (ISO)
    m = re.match(r'^(\d{4})[\/\-\.](\d{1,2})[\/\-\.](\d{1,2})$', s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    # "30 Mar 2026" or "30 March 2026"
    m = re.match(r'(\d{1,2})\s+([a-zA-Z]+)\s+(\d{4})', s)
    if m:
        mon = MONTH_MAP.get(m.group(2).lower())
        if mon:
            try:
                return date(int(m.group(3)), mon, int(m.group(1)))
            except ValueError:
                pass

    # "Mar 30, 2026"
    m = re.match(r'([a-zA-Z]+)\s+(\d{1,2})[,\s]+(\d{4})', s)
    if m:
        mon = MONTH_MAP.get(m.group(1).lower())
        if mon:
            try:
                return date(int(m.group(3)), mon, int(m.group(2)))
            except ValueError:
                pass

    return None


def find_header_row_index(sheet):
    """Find the row index (0-based) that contains the most parseable dates."""
    best_idx = -1
    best_count = 0
    for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=3, values_only=True)):
        count = sum(1 for cell in row if parse_cell_date(cell))
        if count > best_count:
            best_count = count
            best_idx = r_idx
    return best_idx if best_count > 0 else -1


def find_closest_date_col(date_cols, target_date, tolerance_days=10):
    """Find column whose date is closest to target_date within tolerance."""
    best = None
    best_diff = float("inf")
    for dc in date_cols:
        diff = abs((dc["date"] - target_date).days)
        if diff < best_diff and diff <= tolerance_days:
            best_diff = diff
            best = dc
    return best


def find_previous_month_end_col(date_cols, current_date):
    """Find closest month-end (day >= 25) strictly before current_date."""
    best = None
    best_date = None
    for dc in date_cols:
        d = dc["date"]
        if d < current_date and d.day >= 25:
            if best_date is None or d > best_date:
                best_date = d
                best = dc
    return best


# ─────────────────────────────────────────────
# HYPERLINK PARSING
# ─────────────────────────────────────────────

def extract_sheet_name_from_hyperlink(cell):
    """
    Extract sheet name from Excel HYPERLINK formula or hyperlink object.
    Formula: =HYPERLINK("#SheetName!A1","Display Name")
    OR the cell value may just be the display name and we match by name.
    """
    # Try to get hyperlink from cell
    if hasattr(cell, 'hyperlink') and cell.hyperlink:
        target = cell.hyperlink.target or ""
        # Internal link format: #SheetName!A1
        if target.startswith("#"):
            ref = target[1:]  # remove #
            if "!" in ref:
                return ref.split("!")[0].strip("'")
            return ref.strip("'")

    # Try formula in cell value
    val = str(cell.value or "")
    m = re.search(r'HYPERLINK\s*\(\s*["\']#([^"\'!]+)', val, re.IGNORECASE)
    if m:
        return m.group(1).strip("'").strip()

    # Fallback: cell display value = sheet name
    return str(cell.value).strip() if cell.value else None


# ─────────────────────────────────────────────
# CORE PROCESSING
# ─────────────────────────────────────────────

def get_rank_summary_sheets(wb):
    """Return list of sheet names that look like Rank Summary sheets."""
    results = []
    for name in wb.sheetnames:
        if re.search(r'rank\s*summary', name, re.IGNORECASE):
            results.append(name)
    return results


def process_project_sheet(proj_sheet, current_date):
    """
    Read a project sheet, find the two date columns,
    calculate and return the 6 metrics.
    Returns dict with metrics or {'error': 'message'}.
    """
    # Read all values
    all_rows = list(proj_sheet.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return {"error": "Sheet is empty or has too few rows"}

    # Find header row
    best_idx = -1
    best_count = 0
    for r_idx, row in enumerate(all_rows[:3]):
        count = sum(1 for cell in row if parse_cell_date(cell))
        if count > best_count:
            best_count = count
            best_idx = r_idx

    if best_idx == -1:
        return {"error": "No date columns found in header row"}

    header_row = all_rows[best_idx]

    # Parse all date columns from header
    date_cols = []
    for c_idx, cell_val in enumerate(header_row):
        d = parse_cell_date(cell_val)
        if d:
            date_cols.append({"col_idx": c_idx, "date": d})

    if not date_cols:
        return {"error": "No parseable dates found in header"}

    # Find current and previous columns
    curr_col = find_closest_date_col(date_cols, current_date)
    if not curr_col:
        return {"error": f"No column found matching date ~{current_date.strftime('%d/%m/%Y')}"}

    prev_col = find_previous_month_end_col(date_cols, curr_col["date"])
    if not prev_col:
        return {"error": f"No previous month-end column found before {curr_col['date'].strftime('%d/%m/%Y')}"}

    # Read keyword rows — all rows after header
    keywords = []
    for r_idx in range(best_idx + 1, len(all_rows)):
        row = all_rows[r_idx]

        # Column A = SL No — must be a number to be a keyword row
        sl_val = row[0] if len(row) > 0 else None
        if sl_val is None or str(sl_val).strip() == "":
            continue
        try:
            int(float(str(sl_val).strip()))
        except (ValueError, TypeError):
            continue

        # Column B = keyword — must not be empty
        kw_val = row[1] if len(row) > 1 else None
        if not kw_val or str(kw_val).strip() == "":
            continue

        prev_val = row[prev_col["col_idx"]] if prev_col["col_idx"] < len(row) else None
        curr_val = row[curr_col["col_idx"]] if curr_col["col_idx"] < len(row) else None

        keywords.append({"prev": prev_val, "curr": curr_val})

    if not keywords:
        return {"error": "No keyword rows found"}

    metrics = calculate_metrics(keywords)
    metrics["curr_date"] = curr_col["date"].strftime("%d/%m/%Y")
    metrics["prev_date"] = prev_col["date"].strftime("%d/%m/%Y")
    return metrics


def refresh_summary_sheet(wb, summary_sheet_name, progress_callback=None):
    """
    Main processing function.
    Reads the summary sheet, processes each project, writes results back.
    Returns (success_count, error_count, error_details, totals).
    """
    summary_sheet = wb[summary_sheet_name]

    # Parse current date from sheet name
    current_date = parse_date_from_sheet_name(summary_sheet_name)
    if not current_date:
        return 0, 0, [f"Could not parse date from sheet name: '{summary_sheet_name}'"], {}

    # Build sheet lookup by name
    sheet_by_name = {s.lower(): wb[s] for s in wb.sheetnames}

    # Read all rows from summary sheet
    all_rows = list(summary_sheet.iter_rows(min_row=2))  # skip header row 1

    success_count = 0
    error_count   = 0
    error_details = []
    totals = {"total": 0, "first_page": 0, "rank_up": 0, "rank_down": 0, "top5": 0, "top3": 0}

    project_rows = []
    for row in all_rows:
        cell_b = row[1] if len(row) > 1 else None  # Column B
        if cell_b is None or cell_b.value is None:
            continue
        project_rows.append(row)

    total_projects = len(project_rows)

    for idx, row in enumerate(project_rows):
        cell_b = row[1]  # Column B — project name / hyperlink

        # Get project sheet name from hyperlink or cell value
        proj_sheet_name = extract_sheet_name_from_hyperlink(cell_b)
        display_name    = str(cell_b.value).strip() if cell_b.value else f"Row {row[0].row}"

        if progress_callback:
            progress_callback(idx, total_projects, display_name)

        # Find the project sheet
        proj_sheet = None
        if proj_sheet_name:
            proj_sheet = sheet_by_name.get(proj_sheet_name.lower())

        # Fallback: try matching display name
        if not proj_sheet and display_name:
            proj_sheet = sheet_by_name.get(display_name.lower())

        if not proj_sheet:
            error_msg = f"Tab not found: '{proj_sheet_name or display_name}'"
            _write_error_row(summary_sheet, row)
            error_count += 1
            error_details.append(f"Row {row[0].row} — {display_name}: {error_msg}")
            continue

        # Process the project sheet
        result = process_project_sheet(proj_sheet, current_date)

        if "error" in result and result["error"]:
            _write_error_row(summary_sheet, row)
            error_count += 1
            error_details.append(f"Row {row[0].row} — {display_name}: {result['error']}")
        else:
            # Write metrics to columns D(4), E(5), F(6), G(7), H(8), I(9)
            row[3].value = result["total"]       # D
            row[4].value = result["first_page"]  # E
            row[5].value = result["rank_up"]     # F
            row[6].value = result["rank_down"]   # G
            row[7].value = result["top5"]        # H
            row[8].value = result["top3"]        # I
            success_count += 1

            # Accumulate totals
            for k in totals:
                totals[k] += result.get(k, 0)

    return success_count, error_count, error_details, totals


def _write_error_row(summary_sheet, row):
    """Write 'Error' to all 6 metric cells for a failed project row."""
    for col_offset in range(3, 9):  # D=3, E=4, F=5, G=6, H=7, I=8 (0-indexed)
        if col_offset < len(row):
            row[col_offset].value = "Error"


# ─────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────

def main():
    # Header
    st.markdown("""
    <div class="app-header">
        <div class="header-icon">📊</div>
        <div>
            <p class="header-title">OSAM SEO Rank Refresher</p>
            <p class="header-sub">Automated rank summary calculator · OneDrive sync</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    config = get_config()

    # ── Step 1: Connect & Load File ──────────────────────────
    st.markdown('<div class="card"><div class="card-title">Step 1 — Connect to OneDrive</div>', unsafe_allow_html=True)

    if "workbook_bytes" not in st.session_state:
        st.session_state.workbook_bytes = None
        st.session_state.workbook       = None
        st.session_state.token          = None
        st.session_state.sheets_list    = []
        st.session_state.connected      = False

    if not st.session_state.connected:
        if st.button("🔗 Connect to OneDrive & Load File"):
            with st.spinner("Connecting to OneDrive…"):
                try:
                    token = get_access_token(
                        config["client_id"],
                        config["client_secret"],
                        config["tenant_id"]
                    )
                    st.session_state.token = token

                    wb_bytes = download_workbook(token, config["file_res_id"])
                    st.session_state.workbook_bytes = wb_bytes.getvalue()

                    wb = openpyxl.load_workbook(BytesIO(st.session_state.workbook_bytes))
                    st.session_state.workbook    = wb
                    st.session_state.sheets_list = get_rank_summary_sheets(wb)
                    st.session_state.connected   = True

                    st.markdown(f"""
                    <div class="status-box status-ok">
                        ✅ Connected! Found <strong>{len(wb.sheetnames)}</strong> sheets,
                        <strong>{len(st.session_state.sheets_list)}</strong> Rank Summary sheet(s).
                    </div>""", unsafe_allow_html=True)

                except Exception as e:
                    st.markdown(f"""
                    <div class="status-box status-error">
                        ❌ Connection failed: {str(e)}<br>
                        Please check your secrets configuration.
                    </div>""", unsafe_allow_html=True)
    else:
        wb = st.session_state.workbook
        st.markdown(f"""
        <div class="status-box status-ok">
            ✅ Connected · <strong>{len(wb.sheetnames)}</strong> sheets loaded ·
            {len(st.session_state.sheets_list)} Rank Summary sheet(s) found
        </div>""", unsafe_allow_html=True)

        if st.button("🔄 Reconnect / Reload File"):
            for k in ["workbook_bytes","workbook","token","sheets_list","connected"]:
                if k in st.session_state: del st.session_state[k]
            st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Step 2: Select Sheet ─────────────────────────────────
    if st.session_state.connected and st.session_state.sheets_list:
        st.markdown('<div class="card"><div class="card-title">Step 2 — Select Rank Summary Sheet</div>', unsafe_allow_html=True)

        selected_sheet = st.selectbox(
            "Choose which Rank Summary to refresh:",
            options=st.session_state.sheets_list,
            label_visibility="collapsed"
        )

        # Show detected date
        detected_date = parse_date_from_sheet_name(selected_sheet)
        if detected_date:
            # Count project rows
            wb = st.session_state.workbook
            summary_ws = wb[selected_sheet]
            project_count = sum(
                1 for row in summary_ws.iter_rows(min_row=2, values_only=True)
                if row[1] is not None and str(row[1]).strip() != ""
            )
            st.markdown(f"""
            <div class="status-box status-info">
                📅 Detected date: <strong>{detected_date.strftime('%d %B %Y')}</strong> &nbsp;·&nbsp;
                🗂 Projects found: <strong>{project_count}</strong>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="status-box status-warn">
                ⚠️ Could not detect a date from this sheet name.
                Make sure it contains a month and day (e.g. "March 31st Rank Summary").
            </div>""", unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

        # ── Step 3: Refresh ───────────────────────────────────
        st.markdown('<div class="card"><div class="card-title">Step 3 — Refresh</div>', unsafe_allow_html=True)

        if detected_date:
            if st.button(f"🚀 Refresh '{selected_sheet}'"):
                wb = openpyxl.load_workbook(BytesIO(st.session_state.workbook_bytes))

                progress_bar  = st.progress(0)
                status_text   = st.empty()
                log_container = st.empty()
                log_lines     = []

                def progress_callback(idx, total, project_name):
                    pct = int((idx / max(total, 1)) * 100)
                    progress_bar.progress(pct)
                    status_text.markdown(
                        f'<div class="status-box status-info">⏳ Processing {idx+1}/{total}: <strong>{project_name}</strong></div>',
                        unsafe_allow_html=True
                    )
                    log_lines.append(f"✓ {project_name}")
                    log_container.text("\n".join(log_lines[-8:]))

                with st.spinner("Calculating metrics…"):
                    try:
                        success, errors, error_details, totals = refresh_summary_sheet(
                            wb, selected_sheet, progress_callback
                        )

                        progress_bar.progress(100)
                        status_text.empty()
                        log_container.empty()

                        # Save workbook back to bytes
                        out = BytesIO()
                        wb.save(out)
                        updated_bytes = out.getvalue()

                        # Upload back to OneDrive
                        status_text.markdown(
                            '<div class="status-box status-info">☁️ Uploading to OneDrive…</div>',
                            unsafe_allow_html=True
                        )
                        upload_workbook(
                            st.session_state.token,
                            config["file_res_id"],
                            updated_bytes
                        )
                        status_text.empty()

                        # Update session state
                        st.session_state.workbook_bytes = updated_bytes
                        st.session_state.workbook = openpyxl.load_workbook(BytesIO(updated_bytes))

                        # Success summary
                        total_projects = success + errors
                        st.markdown(f"""
                        <div class="status-box status-ok">
                            ✅ Done! <strong>{success}/{total_projects}</strong> projects updated successfully.
                            {f'<br>⚠️ {errors} project(s) had errors — check details below.' if errors > 0 else ''}
                        </div>""", unsafe_allow_html=True)

                        # Metrics summary
                        if success > 0:
                            st.markdown("""
                            <div class="card">
                            <div class="card-title">Summary Totals Across All Projects</div>
                            <div class="metric-grid">
                            """, unsafe_allow_html=True)

                            metrics_display = [
                                ("Total Keywords", totals["total"],      "#4b9fff"),
                                ("1st Page",        totals["first_page"], "#818cf8"),
                                ("Rank Up ↑",       totals["rank_up"],    "#34d399"),
                                ("Rank Down ↓",     totals["rank_down"],  "#f87171"),
                                ("Top 5",           totals["top5"],       "#60a5fa"),
                                ("Top 3",           totals["top3"],       "#fbbf24"),
                            ]

                            cols = st.columns(3)
                            for i, (label, val, color) in enumerate(metrics_display):
                                with cols[i % 3]:
                                    st.markdown(f"""
                                    <div class="metric-card">
                                        <div class="metric-value" style="color:{color}">{val}</div>
                                        <div class="metric-label">{label}</div>
                                    </div>""", unsafe_allow_html=True)

                            st.markdown('</div></div>', unsafe_allow_html=True)

                        # Error details
                        if error_details:
                            with st.expander(f"⚠️ {errors} Error(s) — click to expand"):
                                for ed in error_details:
                                    st.markdown(f"- {ed}")

                    except Exception as e:
                        status_text.empty()
                        st.markdown(f"""
                        <div class="status-box status-error">
                            ❌ Refresh failed: {str(e)}
                        </div>""", unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

    elif st.session_state.connected and not st.session_state.sheets_list:
        st.markdown("""
        <div class="status-box status-warn">
            ⚠️ No Rank Summary sheets found in this workbook.
            Make sure at least one sheet name contains "Rank Summary".
        </div>""", unsafe_allow_html=True)

    # Footer
    st.markdown("""
    <div style="text-align:center; color:#374151; font-size:12px; margin-top:40px; padding-top:20px; border-top:1px solid #1e2332;">
        OSAM SEO Rank Refresher · Built for internal use · Powered by Microsoft Graph API
    </div>""", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
